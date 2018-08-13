# -*- coding: UTF-8 -*-
from lxml import etree
import requests
import string
from  openpyxl import Workbook
import time
# 脚本之家python页面url
baseUrl = 'http://www.jb51.net/list/list_97_1.htm'
k = list()
n = 0
urlDict = {'python': '97', 'java': '207', 'ajax相关': '5', 'javascript': '3', 'dos/bat': '106',
           'asp.net': '21', 'php编程': '15', '正则表达式': '6', 'asp编程':'2', 'jsp编程': '83',
           '编程10000问':'123','css/html': '4','flex': '237', '脚本加解密': '60', 'web2.0': '62',
           'xml/rss':'64','网页编辑器':'176','hta':'110','vbs':'114','htc':'111','perl': '125',
           '游戏相关':'138','vba':'161','远程脚本':'163','coldfusion':'178', 'ruby专题': '201',
           'autoit':'213','seraphzone':'214','powershell':'234','linux shell':'235','lua':'245',
           'golang':'246','mssql':'113','mysql':'112','mariadb':'252','oracle':'154','db2':'155',
           'mssql2008':'236','mssql2005':'210','sqlite':'215','postgresql':'224','mongodb':'239',
           'redis':'242','access':'105','数据库文摘':'132','数据库其他':'133'}
print("可检索的字段有:",end='')
for i in urlDict.keys():
    n += 1
    if(n % 10 == 1):
        print('\n'+i,end='   ')
    else:
        print(i,end='   ')
rsult = input("\n请输入要搜索的字段>>>")
if(rsult in urlDict):
    u =  urlDict[rsult.lower()]
else:
    print("对不起，未检索到您所需要的字段！！！")
pageNum = 1  # 用来计算爬取的条数
wb = Workbook()
ws = wb.active
ws.title = '脚本之家'+rsult.replace('/','')+'专栏'
ws.cell(row=1, column=1).value = '标题'
ws.cell(row=1, column=2).value = '链接'
ws.cell(row=1, column=3).value = '日期'


def getHtml(url):
    try:
        r = requests.get(url, timeout=30)
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        return r.text
    except:
        return ""


def etreeHtml(html):
    global pageNum
    print("-------------------------------------------------")
    html = etree.HTML(html)
    # 因为每页有四十页
    for page in range(1, 41):
        # 标题
        title = html.xpath('//*[@class="artlist clearfix"]/dl/dt[%s]/a/text()' % page)
        # 日期
        timeData = html.xpath('//*[@class="artlist clearfix"]/dl/dt[%s]/span/text()' % page)
        # 链接(因为用的是相对链接，所以要加上：http://www.jb51.net)
        nextUrl = html.xpath('//*[@class="artlist clearfix"]/dl/dt[%s]/a/@href' % page)
        print('名称: ' + str(title[0]))
        # 打印日期
        print('日期: ' + str(timeData[0]).split(':')[1])
        nextUrl = 'http://www.jb51.net' + nextUrl[0]
        print('url：' + str(nextUrl))
        # ver_info = list(zip(title[0],nextUrl,timeData[0]))
        pageNum = pageNum + 1
        ws.cell(row=pageNum, column=1).value = title[0]
        ws.cell(row=pageNum, column=2).value = nextUrl
        ws.cell(row=pageNum, column=3).value = str(timeData[0]).split(':')[1]



def main():
    startpage = 1
    endpage = 2
    try:
        for i in range(startpage, endpage):
            url = 'http://www.jb51.net/list/list_'+u+'_%s.htm' % i
            print("-------------------------------------------------")
            print('第'+ str(i) +'页url:' + url)
            html = getHtml(url)
            etreeHtml(html)
    except:
        print('')
    finally:
        wb.save('脚本之家'+rsult.replace('/','')+'专栏' + '.xlsx')

if(rsult in urlDict):
    starttime = time.time();
    main()
    endtime = time.time();
    totaltime = endtime - starttime;
    print("爬取完成，耗时：{0:.5f}秒".format(totaltime));
